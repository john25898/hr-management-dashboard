"""
Generate enriched employee data JSON from the Excel file.
Detects departed employees via red font, fixes designations,
and separates employees into proper categories.
"""
import openpyxl
import json
import re
from datetime import datetime as dt_lib, timedelta

EXCEL_PATH = r'C:\Users\ADMIN\Documents\program HR\UJTP HRH Master DATABASE 2025 EM (2).xlsx'
OUTPUT_PATH = r'C:\Users\ADMIN\Documents\program HR\data-visualization-dashboard\data\employees-enriched.json'

wb = openpyxl.load_workbook(EXCEL_PATH)
ws_all = wb['All']
ws_lw = wb['Layworkers']

# ── Designation normalisation map ──
# Fixes variations of the same designation
DESIGNATION_NORMALIZE = {
    'hts counselor': 'HTS Counselor',
    'hts counsellor': 'HTS Counselor',
    'hts provider': 'HTS Counselor',
    'hrio': 'HRIO',
    'hrio - mentor': 'HRIO Mentor',
    'health records information officer': 'HRIO',
    'ict officer': 'ICT Officer',
    'pharmaceutical technologist': 'Pharmaceutical Technologist',
    'pharmacy technologist - supply chain': 'Pharmaceutical Technologist - Supply Chain',
    'pharmaceutical  technologist': 'Pharmaceutical Technologist',
    'pharmaceutical  technologist - supply chain': 'Pharmaceutical Technologist - Supply Chain',
    'pharmaceutical technologist - supply chain': 'Pharmaceutical Technologist - Supply Chain',
    'pharmacy technologist  - supply chain': 'Pharmaceutical Technologist - Supply Chain',
    'counselling psychologist': 'Counseling Psychologist',
    'clinical officer - mentor': 'Clinical Officer - Mentor',
    'mentor - prevention': 'HTS Counselor',  # Zipporah fix
    'mentor-prevention': 'HTS Counselor',
    'county clinicians  (mentor)': 'Clinical Officer - Mentor',
    'county nurse (mentor)': 'County Nurse (Mentor)',
    'm&e assistant': 'M&E Assistant',
    'social worker': 'Social Worker',
    'nurse': 'Nurse',
    'medical laboratory technologist': 'Medical Laboratory Technologist',
}

# ── Designation group mapping for filters ──
# Groups that should appear as merged in the filter dropdown
DESIGNATION_GROUPS = {
    'HRIO': ['HRIO', 'HRIO Mentor', 'Health Records Information Officer'],
    'Clinical Officer': ['Clinical Officer', 'Clinical Officer - Mentor', 'County Clinicians  (mentor)', 'County Clinicians (mentor)'],
}

def normalize_designation(desig):
    if not desig:
        return 'Unknown'
    key = str(desig).strip().lower()
    # Normalize multiple spaces
    key = re.sub(r'\s+', ' ', key)
    # Also normalize keys in the map to handle space variations
    norm_map = {re.sub(r'\s+', ' ', k): v for k, v in DESIGNATION_NORMALIZE.items()}
    if key in norm_map:
        return norm_map[key]
    return str(desig).strip()

def compute_age(dob_str):
    """Compute age from DOB string (ISO or ordinal). Returns integer age or None."""
    if not dob_str:
        return None
    try:
        # If it's already an ISO date
        if '-' in str(dob_str):
            d = dt_lib.strptime(str(dob_str), '%Y-%m-%d')
        else:
            # Try parsing as ordinal
            parsed = parse_ordinal_date(str(dob_str))
            if not parsed:
                return None
            d = dt_lib.strptime(parsed, '%Y-%m-%d')
        today = dt_lib.today()
        age = today.year - d.year - ((today.month, today.day) < (d.month, d.day))
        return age if 0 <= age <= 120 else None
    except:
        return None

def get_designation_group(desig):
    """Map a designation to its filter group name."""
    norm = normalize_designation(desig)
    for group, members in DESIGNATION_GROUPS.items():
        # Check if the normalized designation matches any member
        for member in members:
            if normalize_designation(member) == norm:
                return group
            # Also check if the original matches (for already-normalized values)
            if member.lower() == norm.lower():
                return group
    return norm  # Return the normalized individual designation

# ── Regulatory Body normalisation map ──
# Normalises variations of the same professional council name
REGULATORY_BODY_NORMALIZE = {
    # Clinical Officers Council
    'clinical officers council': 'Clinical Officers Council (COC)',
    'clinical officers council (coc)': 'Clinical Officers Council (COC)',
    'council of clinical officers (clinical officers council (coc))': 'Clinical Officers Council (COC)',
    # Nursing Council of Kenya
    'nursing council of kenya': 'Nursing Council of Kenya (NCK)',
    'nursing council of kenya (nck)': 'Nursing Council of Kenya (NCK)',
    'nck': 'Nursing Council of Kenya (NCK)',
    # Health Records and Information Managers Board
    'health records and information managers board': 'Health Records and Information Managers Board',
    'health records and information managers board (health records and information managers board)': 'Health Records and Information Managers Board',
    'health record and information managers board': 'Health Records and Information Managers Board',
    'health records and information manager board': 'Health Records and Information Managers Board',
    # Pharmacy & Poisons Board
    'pharmacy &amp; poisons board (ppb)': 'Pharmacy & Poisons Board (PPB)',
    'pharmacy & poisons board (ppb)': 'Pharmacy & Poisons Board (PPB)',
    'pharmacy and poisons board': 'Pharmacy & Poisons Board (PPB)',
    'pharmacy and poisons board ((ppb)': 'Pharmacy & Poisons Board (PPB)',
    'ppb': 'Pharmacy & Poisons Board (PPB)',
    # Kenya Medical Laboratory Technicians & Technologists Board
    'kenya medical laboratory technicians &amp; technologists board': 'Kenya Medical Laboratory Technicians & Technologists Board (KMLTTB)',
    'kenya medical laboratory technicians & technologists board': 'Kenya Medical Laboratory Technicians & Technologists Board (KMLTTB)',
    'kenya medical laboratory technicians and technologists board': 'Kenya Medical Laboratory Technicians & Technologists Board (KMLTTB)',
    'kenya medical laboratory technicians and technologists board (kmlttb)': 'Kenya Medical Laboratory Technicians & Technologists Board (KMLTTB)',
    'kenya medical laboratory technicians and technologists board (kenya medical laboratory technicians and technologists board (kmlttb))': 'Kenya Medical Laboratory Technicians & Technologists Board (KMLTTB)',
}

def normalize_regulatory_body(val):
    """Normalize regulatory body name to a canonical form."""
    if not val:
        return None
    key = str(val).strip().lower()
    key = re.sub(r'\s+', ' ', key)
    return REGULATORY_BODY_NORMALIZE.get(key, str(val).strip())

def is_departed(cell):
    """Check if a cell has red font indicating departed status."""
    if cell.font and cell.font.color:
        try:
            fc = str(cell.font.color.rgb).upper()
            if 'FF0000' in fc:
                return True
        except:
            pass
    return False

def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

# Month name mapping for ordinal date parsing
MONTH_MAP = {
    'jan': 1, 'january': 1,
    'feb': 2, 'february': 2, 'febr': 2,
    'mar': 3, 'march': 3,
    'apr': 4, 'april': 4,
    'may': 5,
    'jun': 6, 'june': 6,
    'jul': 7, 'july': 7,
    'aug': 8, 'august': 8,
    'sep': 9, 'sept': 9, 'september': 9,
    'oct': 10, 'october': 10,
    'nov': 11, 'november': 11,
    'dec': 12, 'december': 12,
}

def parse_ordinal_date(s):
    """Parse ordinal date strings like '1st May 2021', '1ST OCT 2021', '2nd  Jan 2025' etc."""
    if not s:
        return None
    s = str(s).strip()
    # Collapse multiple spaces
    s = re.sub(r'\s+', ' ', s)
    # Pattern: day(st|nd|rd|th) month_name[,] year  (comma optional, spaces flexible)
    m = re.match(r'(\d+)(?:st|nd|rd|th)?\s*([a-zA-Z]+)\s*,?\s*(\d{4})', s, re.IGNORECASE)
    if not m:
        return None
    day = int(m.group(1))
    month_str = m.group(2).strip().lower()
    year = int(m.group(3))
    month = MONTH_MAP.get(month_str)
    if month is None or day < 1 or day > 31:
        return None
    try:
        return dt_lib(year, month, day).strftime('%Y-%m-%d')
    except:
        return None

def serial_date_to_str(val):
    """Convert Excel serial date, datetime, or ordinal date string to ISO string."""
    if val is None:
        return None
    if isinstance(val, dt_lib):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)):
        try:
            base = dt_lib(1899, 12, 30)
            d = base + timedelta(days=val)
            return d.strftime('%Y-%m-%d')
        except:
            return str(val)
    # String — try ordinal date first, then return as-is
    s = str(val).strip()
    parsed = parse_ordinal_date(s)
    if parsed:
        return parsed
    # Must be a license code or other non-date string
    return s if s else None

# Custom JSON encoder to handle any remaining datetime/non-serializable types
class CustomEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (dt_lib,)):
            return obj.strftime('%Y-%m-%d')
        try:
            return super().default(obj)
        except:
            return str(obj)

# ══════════════════════════════════════
# 1) Process All sheet
# ══════════════════════════════════════
all_employees = []
departed_employees = []

for r in range(3, ws_all.max_row + 1):
    name = ws_all.cell(row=r, column=2).value
    if not name:
        continue
    name = str(name).strip()
    if not name or name.upper() == 'EMPLOYEE NAME':
        continue
    
    cell_b = ws_all.cell(row=r, column=2)
    departed = is_departed(cell_b)
    
    raw_desig = ws_all.cell(row=r, column=7).value
    designation = normalize_designation(raw_desig)
    designation_group = get_designation_group(raw_desig)
    
    emp = {
        'name': name,
        'gender': safe_str(ws_all.cell(row=r, column=3).value),
        'phone': safe_str(ws_all.cell(row=r, column=4).value),
        'idNo': safe_str(ws_all.cell(row=r, column=5).value),
        'designation': designation,
        'designationGroup': designation_group,
        'designationOriginal': safe_str(raw_desig),
        'county': safe_str(ws_all.cell(row=r, column=8).value),
        'subCounty': safe_str(ws_all.cell(row=r, column=9).value),
        'station': safe_str(ws_all.cell(row=r, column=10).value),
        'dateEmployed': serial_date_to_str(ws_all.cell(row=r, column=11).value),  # DATE EMPLOYED
        'contractEnd': serial_date_to_str(ws_all.cell(row=r, column=12).value),   # contract End date (datetime!)
        'dob': serial_date_to_str(ws_all.cell(row=r, column=13).value),           # DOB (ordinal strings)
        'educationLevel': safe_str(ws_all.cell(row=r, column=16).value),          # EDUCATION LEVEL
        'qualification': safe_str(ws_all.cell(row=r, column=17).value),           # QUALIFICATION
        'othersCert': safe_str(ws_all.cell(row=r, column=19).value),              # others (Nascop cert)
        'regulatoryBody': normalize_regulatory_body(ws_all.cell(row=r, column=20).value),  # REGULATORY BODY
        'practiseeLicence': safe_str(ws_all.cell(row=r, column=21).value),        # PRACTISING LICENCE (license ref)
        'validUntil': serial_date_to_str(ws_all.cell(row=r, column=22).value),    # VALID UNTIL (actual dates!)
        'age': compute_age(safe_str(ws_all.cell(row=r, column=13).value)),        # Computed from DOB
        'isDeparted': departed,
    }
    
    if departed:
        departed_employees.append(emp)
    else:
        all_employees.append(emp)

# ══════════════════════════════════════
# 2) Process Layworkers sheet
# ══════════════════════════════════════
layworkers = []
for r in range(2, ws_lw.max_row + 1):
    name = ws_lw.cell(row=r, column=2).value
    if not name:
        continue
    name = str(name).strip()
    if not name:
        continue
    
    lw = {
        'id': ws_lw.cell(row=r, column=1).value,
        'name': name,
        'facility': safe_str(ws_lw.cell(row=r, column=3).value),
        'county': safe_str(ws_lw.cell(row=r, column=4).value),
        'subCounty': safe_str(ws_lw.cell(row=r, column=5).value),
        'cadre': safe_str(ws_lw.cell(row=r, column=6).value),
        'phone': safe_str(ws_lw.cell(row=r, column=7).value),
        'idNo': safe_str(ws_lw.cell(row=r, column=8).value),
        'amount': ws_lw.cell(row=r, column=9).value,
    }
    layworkers.append(lw)

# ══════════════════════════════════════
# 3) Designation groups for filters
# ══════════════════════════════════════
# Build the merged filter list
all_designation_groups = {}
for emp in all_employees:
    grp = emp['designationGroup']
    all_designation_groups[grp] = all_designation_groups.get(grp, 0) + 1

merged_filter_designations = sorted(all_designation_groups.keys())

# Also build individual designation list for reference
individual_designations = sorted(set(emp['designation'] for emp in all_employees))

# ══════════════════════════════════════
# 4) Summary
# ══════════════════════════════════════
summary = {
    'totalEmployees': len(all_employees),
    'totalDeparted': len(departed_employees),
    'totalLayworkers': len(layworkers),
    'totalPersonnel': len(all_employees) + len(layworkers),
    'designationGroups': merged_filter_designations,
    'individualDesignations': individual_designations,
}

# ══════════════════════════════════════
# 5) Write output
# ══════════════════════════════════════
output = {
    'employees': all_employees,
    'departed': departed_employees,
    'layworkers': layworkers,
    'summary': summary,
}

with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, ensure_ascii=False, cls=CustomEncoder)

print(f"✅ Data written to: {OUTPUT_PATH}")
print(f"\n📊 Summary:")
print(f"   Active employees: {len(all_employees)}")
print(f"   Departed employees: {len(departed_employees)}")
print(f"   Layworkers: {len(layworkers)}")
print(f"   Total personnel: {len(all_employees) + len(layworkers)}")
print(f"\n🏷️  Designation groups for filters ({len(merged_filter_designations)}):")
for g in merged_filter_designations:
    print(f"   - {g}: {all_designation_groups[g]}")
print(f"\n🔍 Individual designations ({len(individual_designations)}):")
for d in individual_designations:
    count = sum(1 for e in all_employees if e['designation'] == d)
    print(f"   - {d}: {count}")

# Verify Zipporah fix
zipporah = [e for e in all_employees if 'zipporah' in e['name'].lower()]
if zipporah:
    print(f"\n✅ Zipporah fix verified: {zipporah[0]['name']} → {zipporah[0]['designation']}")
